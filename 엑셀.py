from openpyxl import load_workbook

# data_only=Ture로 해줘야 수식이 아닌 값으로 받아온다.
load_wb = load_workbook("Test.xlsx", data_only=True)
# 시트 이름으로 불러오기
load_ws = load_wb['Sheet1']



# 셀 주소로 값 출력
print(load_ws['A1'].value)

# 셀 좌표로 값 출력
a = load_ws.cell(1, 2).value
b = "바보",a
print(b)
print(load_ws.cell(1, 2).value)
